import tkinter as tk
from tkinter import messagebox, ttk
import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

class MainApp:
    def __init__(self, master):
        self.master = master
        master.title("Main Menu")
        master.geometry("500x500")

        self.background_color = "#252521"
        self.button_color = "#c5440a"
        self.text_color = "#000000"

        self.master.configure(bg=self.background_color)

        # Get the directory of the Python script
        script_dir = os.path.dirname(os.path.abspath(__file__))

        # Construct the path to the logo image file
        logo_path = os.path.join(script_dir, "logo.png")

        # Use the constructed path to create the PhotoImage
        self.logo = tk.PhotoImage(file=logo_path)

        # Place logo in the lower right corner
        self.logo_label = tk.Label(master, image=self.logo, bg=self.background_color)
        self.logo_label.place(relx=1.0, rely=1.0, anchor="se")

        # Menu buttons
        self.register_button = tk.Button(master, text="Registrera användare", bg=self.button_color, fg=self.text_color,
                                         command=self.open_registration_window)
        self.register_button.place(relx=0.5, rely=0.3, anchor="center", width=200)

        self.log_button = tk.Button(master, text="Logga besök", bg=self.button_color, fg=self.text_color,
                                    command=self.open_logging_window)
        self.log_button.place(relx=0.5, rely=0.7, anchor="center", width=200)

    def open_registration_window(self):
        registration_window = tk.Toplevel(self.master)
        registration_window.title("Registrera användare")
        registration_window.geometry("500x500")
        registration_window.configure(bg=self.background_color)
        registration_window.grab_set()  # Make this window modal
        registration_window.focus_set()  # Focus on this window

        UserRegistrationApp(registration_window)

    def open_logging_window(self):
        logging_window = tk.Toplevel(self.master)
        logging_window.title("Logga besök")
        logging_window.geometry("500x500")
        logging_window.configure(bg=self.background_color)
        logging_window.grab_set()  # Make this window modal
        logging_window.focus_set()  # Focus on this window

        DataLoggingApp(logging_window)


class UserRegistrationApp:
    def __init__(self, master):
        self.master = master
        master.geometry("500x500")
        master.configure(bg="#252521")

        # Create a frame to hold the widgets
        self.frame = tk.Frame(master, bg="#252521")
        self.frame.pack(expand=True, fill="both")

        # Load the logo image
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
        self.logo = tk.PhotoImage(file=logo_path)

        # Display the logo
        self.logo_label = tk.Label(master, image=self.logo, bg="#252521")
        self.logo_label.place(relx=1.0, rely=1.0, anchor="se")

        self.label_nick = tk.Label(self.frame, text="Nick:")
        self.label_nick.grid(row=0, column=0, pady=5, padx=5)
        self.entry_nick = tk.Entry(self.frame)
        self.entry_nick.grid(row=0, column=1, pady=5)

        self.label_födelseår = tk.Label(self.frame, text="Födelseår:")
        self.label_födelseår.grid(row=1, column=0, pady=5, padx=5)
        self.entry_födelseår = tk.Entry(self.frame)
        self.entry_födelseår.grid(row=1, column=1, pady=5)

        self.label_kön = tk.Label(self.frame, text="Könsidentitet:")
        self.label_kön.grid(row=2, column=0, pady=5, padx=5)
        self.entry_kön = tk.Entry(self.frame)
        self.entry_kön.grid(row=2, column=1, pady=5)

        self.submit_button = tk.Button(self.frame, text="Registrera användare", command=self.register_user)
        self.submit_button.grid(row=3, column=0, columnspan=2, pady=5)

        # Check if users.json exists, if not, create it
        if not os.path.exists("users.json"):
            with open("users.json", "w") as f:
                json.dump([], f)

        # Check if data.xlsx exists, if not, create it
        if not os.path.exists("data.xlsx"):
            wb = Workbook()
            wb.save("data.xlsx")

        master.bind("<Return>", self.register_user)

        # Center the frame vertically and horizontally
        self.frame.pack_propagate(False)
        self.frame.place(relx=0.5, rely=0.5, anchor="center")

    def register_user(self, event=None):
        nick = self.entry_nick.get()
        födelseår = self.entry_födelseår.get()
        kön = self.entry_kön.get()

        if nick and födelseår and kön:
            if self.is_nick_available(nick):
                user_data = {"Nick": nick, "Födelseår": födelseår, "Kön": kön}
                self.save_user_data(user_data)
                messagebox.showinfo("Success", "Användare registrerad!")
                self.clear_entries()
            else:
                messagebox.showerror("Error", "Nicket är redan taget. Välj ett nytt.")
        else:
            messagebox.showerror("Error", "Fyll i alla fält.")
    def is_nick_available(self, nick):
        with open("users.json", "r") as f:
            data = json.load(f)
            for user in data:
                if user["Nick"] == nick:
                    return False
        return True

    def save_user_data(self, user_data):
        if not os.path.exists("users.json"):
            with open("users.json", "w") as f:
                json.dump([], f)

        with open("users.json", "r+") as f:
            data = json.load(f)
            data.append(user_data)
            f.seek(0)
            json.dump(data, f)

    def clear_entries(self):
        self.entry_nick.delete(0, tk.END)
        self.entry_födelseår.delete(0, tk.END)
        self.entry_kön.delete(0, tk.END)
        


class DataLoggingApp:
    def __init__(self, master):
        self.master = master
        master.geometry("500x500")
        master.configure(bg="#252521")

         # Create a frame to hold the widgets
        self.frame = tk.Frame(master, bg="#252521")
        self.frame.pack(expand=True, fill="both")

        # Load the logo image
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
        self.logo = tk.PhotoImage(file=logo_path)

        # Display the logo
        self.logo_label = tk.Label(master, image=self.logo, bg="#252521")
        self.logo_label.place(relx=1.0, rely=1.0, anchor="se")

        self.label_username = tk.Label(self.frame, text="Nick:")
        self.label_username.grid(row=0, column=0, pady=5, padx=5)
        self.entry_username = tk.Entry(self.frame)
        self.entry_username.grid(row=0, column=1, pady=5)

        self.log_button = tk.Button(self.frame, text="Logga besök", command=self.log_visit)
        self.log_button.grid(row=1, column=0, columnspan=2, pady=5)
        
        master.bind("<Return>", self.log_visit)

        # Center the frame vertically and horizontally
        self.frame.pack_propagate(False)
        self.frame.place(relx=0.5, rely=0.5, anchor="center")

    def log_visit(self, event=None):
        username = self.entry_username.get()

        if username:
            user_data = self.get_user_data(username)
            if user_data:
                self.write_to_excel(user_data)
                messagebox.showinfo("Success", "Data logged successfully!")
                self.entry_username.delete(0, tk.END)
            else:
                messagebox.showerror("Error", "Användaren hittades inte.")
        else:
            messagebox.showerror("Error", "Ange ett användarnamn.")

    def get_user_data(self, username):
        with open("users.json", "r") as f:
            data = json.load(f)
            for user in data:
                if user["Nick"] == username:
                    return user
        return None

    def write_to_excel(self, user_data):
        today = datetime.now()
        month = today.strftime("%B")
        year = str(today.year)

        if not os.path.exists("data.xlsx"):
            wb = Workbook()
            wb.save("data.xlsx")

        wb = load_workbook("data.xlsx")
        if month not in wb.sheetnames:
            wb.create_sheet(month)
            sheet = wb[month]
            sheet.append(["Date", "Time", "Födelseår", "Kön", "Weekday"])

        sheet = wb[month]
        sheet.append([today.strftime("%x"), today.strftime("%X"), user_data["Födelseår"],
                      user_data["Kön"], today.strftime("%A")])
        wb.save("data.xlsx")


def main():
    root = tk.Tk()
    app = MainApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
